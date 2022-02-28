import numpy as np
import openpyxl as px
import matplotlib.pyplot as plt
import os 
import csv

'''Instructions:
1) Export OpenRocket simulation data to '.csv' format. (See example: 'SSI-IREC-2017_OpenRocket.csv') Be sure to use in, lb, s
2) Create the aerodynamic data file. (See example: SSI-IREC-2017_AeroData.xlsx) a. Over a range of Mach numbers, copy the aerodynamic data using the 'Component Analysis' feature of OpenRocket. Follow the exact column format of the example file. Be sure to follow the specified units. Do not change the sheet name from 'Sheet1'.
3) Update the nozzle exit plane location value (see 'OpenRocket Inputs') in the 'DampingCalcs.m' file. Be sure to follow the specified units.
4) Run 'DampingCalcs.m' and follow the prompts to select the two required files from steps 1-2.'''

def run_sim(csv_file_name,xlsx_file_name, X_ne, min_DR=0.05, max_DR=0.3, file_path='-'):
    '''Using OpenRocket simulation data exported as a CSV and an Aero file as a XSLX, runs dynamic stability calculation and outputs graphs to demonstrate stability coefficients as a function of time. X_ne is distance from tip of nosecone to nozzle exit in inches; min/max_DR are limits on dynamic stability ratio acceptable range; file_path is the directory location of the file, not including the file name.'''
    
    #load current file path if not provided
    if file_path=='-':
        file_path = ask_file()

    
    #Load files
    data = load_csvfile(file_path, csv_file_name)
    data = np.asarray(data)
    aero_data = load_xlsfile(file_path,xlsx_file_name)
    aero_data = np.asarray(aero_data)
                            
    #Constants
    rowoffset = 9
    coloffset = 0
    R_air = 287.1 #J/(kg*K)
    gamma = 1.4                       
    time_ind = 0
    cg_ind = 24
    air_temp_ind = 49
    air_pressure_ind = 50
    vert_velocity_ind = 2
    velocity_ind = 4
    area_ref_ind = 45
    prop_mass_ind = 20
    I_long_ind = 21  
                            
    #Aero data extraction                        
    Mach_ind = aero_data[:,0] #Range of Mach numbers over which aerodynamics are characterized
    CN_alpha_ind = aero_data[:,1] 
    CP_ind = aero_data[:,2]
    CN_alpha_nc_ind = aero_data[:,3]
    CN_alpha_cans_ind = aero_data[:,4] 
    CN_alpha_fins_ind = aero_data[:,5]
    CN_alpha_tc_ind = aero_data[:,6]
    CP_nc_ind = aero_data[:,7] # in
    CP_cans_ind = aero_data[:,8] # in
    CP_fins_ind = aero_data[:,9] # in
    CP_tc_ind = aero_data[:,10] #in   
                            
    #Sim data Extraction
    time = data[:,time_ind]
    CG = data[:,cg_ind]
    air_temp = data[:,air_temp_ind]
    air_pressure = data[:,air_pressure_ind]
    vert_velocity = data[:,vert_velocity_ind]
    velocity = data[:,velocity_ind]
    area_ref = data[:,area_ref_ind]
    prop_mass = data[:,prop_mass_ind]
    I_long = data[:,I_long_ind]  
    burnout_time=time[[i for i,x in enumerate(prop_mass) if x==0][0]]
    apogee_time=time[[i for i,x in enumerate(vert_velocity) if x<=0][1]]
    print "Burnout Time: ", burnout_time
    print "Apogee Time: ", apogee_time
    
    #Conversion
    X_ne = 0.0254*X_ne
    CP_ind = np.multiply(0.0254,CP_ind)
    CP_nc_ind = np.multiply(0.0254,CP_nc_ind)
    CP_cans_ind = np.multiply(0.0254,CP_cans_ind)
    CP_fins_ind = np.multiply(0.0254,CP_fins_ind)
    CP_tc_ind = np.multiply(0.0254,CP_tc_ind)
    CG = np.multiply(0.0254,CG) # convert from in to m
                            
    air_temp = np.divide(np.add(air_temp, 459.7),1.8) # convert from F to K
    air_pressure = np.multiply(1e2,air_pressure) # convert from mbar to Pa
    velocity = np.multiply(0.3048,velocity) # ft/s to m/s
    area_ref = np.multiply(6.452e-4,area_ref) # in^2 to m^2
    prop_mass = np.multiply(0.4536,prop_mass) # lb to kg
    I_long = np.multiply(0.04214,I_long) # lb*ft^2 to kg*m^2
    
    #Calculations
    rho_air = np.divide(air_pressure,(R_air*air_temp))
    prop_m_dot = np.multiply(-1,np.divide(np.diff(prop_mass),np.diff(time)))
    prop_m_dot = np.append(prop_m_dot,[0])
    sound_speed = np.sqrt(np.multiply(gamma,np.multiply(R_air,air_temp)))
    mach = np.divide(velocity,sound_speed)
    L_ref = np.sqrt(np.multiply((4/np.pi),area_ref))
    q = np.multiply(0.5,np.multiply(rho_air,np.power(velocity,2)))
                                            
                                                        
    #Interpolation of Aero file
    CN_alpha = np.interp(mach, Mach_ind, CN_alpha_ind)
    CP = np.interp(mach, Mach_ind, CP_ind)
    CN_alpha_nc = np.interp(mach, Mach_ind, CN_alpha_nc_ind)
    CN_alpha_cans = np.interp(mach, Mach_ind, CN_alpha_cans_ind)
    CN_alpha_fins = np.interp(mach, Mach_ind, CN_alpha_fins_ind)
    CN_alpha_tc = np.interp(mach, Mach_ind, CN_alpha_tc_ind)
    CP_nc = np.interp(mach, Mach_ind, CP_nc_ind)
    CP_cans = np.interp(mach, Mach_ind, CP_cans_ind)
    CP_fins = np.interp(mach, Mach_ind, CP_fins_ind)
    CP_tc = np.interp(mach, Mach_ind, CP_tc_ind)
    
    #Calculations
                            
    C1 = 0.5*np.multiply(rho_air,np.multiply(np.power(velocity,2),np.multiply(area_ref,np.multiply(CN_alpha,np.subtract(CP,CG))))) # Corrective moment coefficient
    C2R = np.multiply(prop_m_dot,np.power((np.subtract(X_ne,CG)),2)) # Propulsive damping coefficient
    C2A = 0.5*np.multiply(rho_air,np.multiply(velocity,np.multiply(area_ref,(np.add(np.multiply(CN_alpha_nc,np.power(np.subtract(CP_nc,CG),2)),np.add(np.multiply(CN_alpha_cans,np.power(np.subtract(CP_cans,CG),2)),np.add(np.multiply(CN_alpha_fins,np.power(np.subtract(CP_fins,CG),2)),np.multiply(CN_alpha_tc,np.power(np.subtract(CP_tc,CG),2))))))))) # Aerodynamic damping coefficient
    C2 = np.add(C2R,C2A) # Damping coefficient
    DR = np.divide(C2,np.multiply(2,np.sqrt(np.multiply(C1,I_long)))) # Damping Ratio
    SM = np.divide(np.subtract(CP,CG),L_ref)    
    NF = np.divide(np.sqrt(np.divide(C1,I_long)),2*np.pi) #Natural frequency 
                            
    DR_valid_ind = [i for i,x in enumerate(DR) if ~np.isnan(x)][-1]
    
    
    #Plotting
    plt.figure(1)
    plt.xlabel('Time (s)')
    plt.ylabel('Damping Ratio $\zeta$')
    plt.axis((0,time[len(time)/2],0,0.5))
    plt.plot(time,DR)
    plt.plot(time[:DR_valid_ind],min_DR*np.ones(DR_valid_ind),'--',label='Min')
    plt.plot(time[:DR_valid_ind],max_DR*np.ones(DR_valid_ind),'--',label='Max')
    plt.plot(np.dot(burnout_time,np.ones(2)),[0,1],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,1],'k-.',label='Apogee',linewidth=0.4)
    plt.legend()
    
    plt.figure(2)
    plt.xlabel('Time (s)')
    plt.ylabel('Stability Margin (cal)')
    plt.plot(burnout_time*np.ones(2),[0,12],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,12],'k-.',label='Apogee',linewidth=0.4)
    plt.axis((0,time[len(time)/2],0,12))
    plt.plot(time[:DR_valid_ind],SM[:DR_valid_ind])
    plt.legend()
    
    plt.figure(3)
    plt.xlabel('Time (s)')
    plt.ylabel('Dynamic Pressure (kPa)')
    plt.plot(burnout_time*np.ones(2),[0,100],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,100],'k-.',label='Apogee',linewidth=0.4)
    plt.plot(time[:DR_valid_ind],np.divide(q[:DR_valid_ind],1e3))
    plt.legend()
    
    plt.figure(4)
    plt.xlabel('Time (s)')
    plt.ylabel('Corrective Moment Coefficient (Nm)')
    plt.plot(time[:DR_valid_ind],C1[:DR_valid_ind])
    plt.plot(burnout_time*np.ones(2),[0,10000],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,10000],'k-.',label='Apogee',linewidth=0.4)
    plt.legend()
    
    plt.figure(5)
    plt.xlabel('Time (s)')
    plt.ylabel('Damping Moment Coefficient')
    plt.plot(time[:DR_valid_ind],C2R[:DR_valid_ind],label='Propulsive')
    plt.plot(time[:DR_valid_ind],C2A[:DR_valid_ind],label='Aerodynamic')                    
    plt.plot(time[:DR_valid_ind],C2[:DR_valid_ind],label='Total') 
    plt.plot(burnout_time*np.ones(2),[0,60],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,60],'k-.',label='Apogee',linewidth=0.4)
    plt.legend()
    
    plt.figure(6)
    plt.xlabel('Time (s)')
    plt.ylabel('Natural Frequency (Hz)')
    plt.plot(time[:DR_valid_ind],NF[:DR_valid_ind])
    plt.plot(burnout_time*np.ones(2),[0,7],'r-.',label='Burnout',linewidth=0.4)
    plt.plot(apogee_time*np.ones(2),[0,7],'k-.',label='Apogee',linewidth=0.4)
    plt.legend()
    

def ask_file():
    '''Load current file location'''
    dir_path = os.path.dirname(os.path.realpath(__file__))
    return dir_path

def load_csvfile(file_path,file_name):
    '''Load a CSV at file_path/file_name, collecting and returning the information in a matrix'''
    data = []
    try:
        with open(file_path+"\\"+file_name, 'r') as f:
            for line in f:
                data_line = line.split(",")
                if len(data_line)>1 and line[0]!='#':
                    data.append([float(i) for i in data_line])
        return data
    except:
        print(file_path+"\\"+file_name)
        print 'CSV file doesn\'t exist'
        return -1

def load_xlsfile(file_path,file_name):
    '''Load a XLSX file , collecting and returning the information in a matrix'''
    W = px.load_workbook(file_path+'\\'+file_name)
    sheet_names = W.sheetnames
    p = W[sheet_names[0]]

    a=[]
    i = 0
    j = 0
    rows = 0
    for row in p.iter_rows():
        if i!= 0:
            for k in row:
                j+=1
                a.append(float(k.internal_value))
        i+=1
        if rows == 0:
            rows = j
    a=[a]
    aa= np.resize(a, (rows, i))
    return aa